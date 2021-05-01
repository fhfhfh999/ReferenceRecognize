# import xlwt
from openpyxl import Workbook
import re


class document:

    def seperate_sentences(self, sentences):
        self.wb = Workbook()

        self.ws = self.wb.active
        self.ws.title = 'sheet1'

        # self.workbook = xlwt.Workbook(encoding="utf-8")
        # self.worksheet = self.workbook.add_sheet()
        row0 = ['序号', '题名', '责任者', '来源名称', '出版年', '卷', '期', '页码', 'Doi']
        for i in range(1, len(row0)+1):
            self.ws.cell(4, i, row0[i-1])
        self.ws.cell(2, 1, 'Google Scholar学者主页网址：')
        for i in range(1, len(sentences) - 1):
            sentence = sentences[i]
            seperated_sentences = analysis(sentence)
            self.ws.cell(i + 5, 1, i)
            self.ws.cell(i + 5, 2, seperated_sentences[1])
            namelist = ""
            for name in seperated_sentences[0]:
                namelist += name
            self.ws.cell(i + 5, 3, namelist)
            self.ws.cell(i + 5, 4, seperated_sentences[2])
        self.wb.save('simple_excel.xlsx')
        file = open("待确认.txt", 'w', encoding="utf-8")
        file.write(sentences[-1])
        file.close()


def analysis(sentence):
    # print(sentence)
    # print('\n')
    return sentence_operate(sentence)


def sentence_operate(sentence):
    begin = 0
    tem = []
    answer = []
    loop = True
    while loop:
        end = sentence.find(",", begin)
        if end - begin > 40:
            break
        element = sentence[begin:end].lstrip()
        if element[:4] == "and ":
            element = element[4:]
            loop = False
        elif re.search(" and ", element) is not None:
            tem = tem + element.split(" and ")
            begin = re.search(" and ", element).span()[1]
            begin = sentence.find(",", begin)
            break

        tem.append(element)
        begin = end + 1
    answer.append(tem)
    if sentence.find("'", begin, begin + 5) != -1:
        sen, begin = apostrophe(sentence, begin)
        answer.append(sen)
    elif sentence.find('"', begin, begin + 5) != -1:
        sen, begin = DoubleQuotes(sentence, begin)
        answer.append(sen)
    else:
        end = sentence.find(",", begin)
        answer.append(sentence[begin:end])
        begin = end + 1
    answer.append(sentence[begin:])
    # print(answer)
    return answer


def apostrophe(sentence, begin):
    ans = re.search("['](.*)[']", sentence[begin:]).span()
    # print(ans)
    end = ans[1] + begin
    return sentence[ans[0]+begin+1:ans[1]+begin-2], end


def DoubleQuotes(sentence, begin):
    ans = re.search('["](.*)["]', sentence[begin:]).span()
    # print(ans)
    end = ans[1] + begin
    return sentence[ans[0]+begin+1:ans[1]+begin-2], end


# def sentence_operate(sentence):
#     # seper = sentence.split(",")
#     seper = re.split(r",(?![^\']*\')", sentence)
#     fin = []
#     i = 0
#     tem = []
#     for element in seper:
#         element = element.lstrip()
#         if i == 0:
#             if element[:4] == "and ":
#                 element = element[4:]
#                 tem.append(element)
#                 fin.append(tem)
#                 i += 1
#                 continue
#             elif len(element) > 30:
#                 fin.append(tem)
#                 i += 1
#             else:
#                 tem.append(element.strip())
#         if i == 1:
#             fin.append(element.strip())
#             i += 1
#             tem = []
#             continue
#         if i == 2:
#             tem.append(element)
#     fin.append(tem)
#     print(fin)


if __name__ == "__main__":
    text = "Z. Liu, D. Chen, and X. Chen, “CpG Island Identification with higher order and variable order Markov Models,” Data Mining in Biomedicine, series: Springer Optimization and Its Applications, vol. 7, Pardalos, Panos M.; Boginski, Vladimir, L.; Vazacopoulos, Alkis (Eds.), Springer, 2007: 47-58."
    sentence_operate(text)
